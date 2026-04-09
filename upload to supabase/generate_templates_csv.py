"""
generate_templates_csv.py — Génère models_final.csv depuis Templates/<Plan>/<Langue>/<Examen>.xlsx
=================================================================================================
Structure attendue :
    Templates/
      Free/   Anglais/   Echographie.xlsx
      Solo/   Francais/  IRM.xlsx
      Pro/    ...
      Clinic/ ...

Dans chaque xlsx :
    - Nom de la feuille  → category
    - Ligne 1 (headers) → name
    - Ligne 2           → content

Colonnes du CSV produit : language, category, exam_type, name, content, plan
Exécution : py -3 generate_templates_csv.py
"""

import os
import glob
import pandas as pd
import openpyxl

# Dossier racine des templates
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "Templates")
OUTPUT_CSV    = os.path.join(os.path.dirname(__file__), "models_final.csv")

# Normalisation du nom de plan (dossier → valeur BDD)
PLAN_MAP = {
    "free":   "free",
    "solo":   "solo",
    "pro":    "pro",
    "clinic": "clinic",
}


def extract_templates_from_xlsx(filepath: str, plan: str, language: str) -> list[dict]:
    """Lit un fichier xlsx et retourne la liste des modèles."""
    exam_type = os.path.splitext(os.path.basename(filepath))[0]
    records   = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ⚠  Impossible d'ouvrir {filepath} : {e}")
        return records

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        category = sheet_name.strip()

        # Ligne 1 = noms, Ligne 2 = contenus
        names    = [cell.value for cell in ws[1]]
        contents = [cell.value for cell in ws[2]]

        for name, content in zip(names, contents):
            if name is None or str(name).strip() == "":
                continue
            records.append({
                "language":  language,
                "category":  category,
                "exam_type": exam_type,
                "name":      str(name).strip(),
                "content":   str(content).strip() if content else "",
                "plan":      plan,
            })

    return records


def main():
    all_records = []
    xlsx_files  = glob.glob(
        os.path.join(TEMPLATES_DIR, "**", "*.xlsx"),
        recursive=True,
    )
    # Exclure les fichiers chiffrés
    xlsx_files = [f for f in xlsx_files if not f.endswith(".enc")]

    print(f"📂 {len(xlsx_files)} fichiers xlsx trouvés dans {TEMPLATES_DIR}\n")

    for filepath in sorted(xlsx_files):
        # Templates/<Plan>/<Langue>/<Examen>.xlsx
        rel   = os.path.relpath(filepath, TEMPLATES_DIR)
        parts = rel.replace("\\", "/").split("/")
        if len(parts) != 3:
            print(f"  ⏭  Ignoré (chemin inattendu) : {rel}")
            continue

        plan_raw, language, _ = parts
        plan = PLAN_MAP.get(plan_raw.lower())
        if plan is None:
            print(f"  ⏭  Plan inconnu '{plan_raw}', ignoré : {rel}")
            continue

        records = extract_templates_from_xlsx(filepath, plan, language)
        if records:
            print(f"  ✅ {plan:7} | {language:15} | {parts[2]:45} → {len(records)} modèles")
        else:
            print(f"  ⚠  {plan:7} | {language:15} | {parts[2]:45} → vide")
        all_records.extend(records)

    if not all_records:
        print("\n❌ Aucun modèle extrait.")
        return

    df = pd.DataFrame(all_records, columns=["language", "category", "exam_type", "name", "content", "plan"])
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    print(f"\n{'='*60}")
    print(f"✅ {len(df)} modèles exportés → {OUTPUT_CSV}")
    print(f"   Plans     : {sorted(df['plan'].unique())}")
    print(f"   Langues   : {len(df['language'].unique())}")
    print(f"   Catégories: {len(df['category'].unique())}")
    print(f"{'='*60}")
    print("\n▶  Lancez ensuite : py -3 supabase_upload.py")


if __name__ == "__main__":
    main()
