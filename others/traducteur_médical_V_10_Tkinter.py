import os
import sys
import time
import json
import re
import random
import threading
import hashlib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import concurrent.futures

# ==========================================
# 1. MOTEUR DE DONNÉES (GLOSSAIRE & CONFIG)
# ==========================================

# Liste ultra-complète des termes à ne jamais traduire (Invariants)
PROTECTED_TERMS = [
    # --- Séquences & Protocoles ---
    "T1", "T2", "T2*", "FLAIR", "STIR", "HASTE", "VIBE", "DIXON", "TRUFISP", "THRIVE",
    "Fat Sat", "FS", "MPR", "MIP", "MinIP", "3D", "2D", "DWI", "SWI", "ADC",
    "3D cosmic", "VIBE FS", "T1 Vibe", "T2 DP", "DP FS", "DTI", "BOLD", "ASL", "ToF", "PC", "SSFP",
    
    # --- Modalités & Dosimétrie ---
    "PET", "PET-CT", "CBCT", "DEXA", "DXA",
    "CTDI", "CTDIvol", "DLP", "SUV", "SUVmax", "SUVmean", 
    
    # --- Produits de contraste & Traceurs ---
    "Gadolinium", "Dotarem", "Clarity", "Omnipaque", "Visipaque", "IV",
    "18F-FDG", "FDG", "Tc99m", "PSMA", "DOPA", "DOTATOC", "F-Choline",
    
    # --- Unités de Mesure ---
    "mm", "cm", "ml", "cc", "kV", "mAs", "bpm", "MHz", "mSv", "mGy", "HU", "cm³", "ng/ml", "ng/ml/cm³", "Bq", "MBq",
    
    # --- Paramètres IRM / Scanner ---
    "1.5T", "3T", "7T", "TR", "TE", "TI", "FA", "FOV", "b0", "b50", "b100", "b400", "b800", "b1000", "b2000",
    
    # --- Scores & Classifications ---
    "PI-RADS", "TI-RADS", "BI-RADS", "CAD-RADS", "LI-RADS", "VI-RADS", "O-RADS", "ACR", "Bosniak", "Gleason", "Lugano", "Deauville",
    
    # --- Niveaux Vertébraux (Invariants internationaux) ---
    "C1", "C2", "C3", "C4", "C5", "C6", "C7",
    "Th1", "Th2", "Th3", "Th4", "Th5", "Th6", "Th7", "Th8", "Th9", "Th10", "Th11", "Th12",
    "L1", "L2", "L3", "L4", "L5", "S1", "C1-C2", "L4-L5", "L5-S1",
    
    # --- Autres ---
    "PSA", "V3", "V4", "TFCC", "Wallis", "PACS", "DICOM", "RIS", "HIS"
]

KEYS_LIST = ["ORL", "ASP", "TAP", "Uroscanner", "Écho", "IRM", "CPRM", "CHC", "SA", "transsonore", "anéchogène", "PDC", "VBP", "TSA", "TDM", "TEP"]

MEDICAL_GLOSSARY = {
    "KEYS": KEYS_LIST,
    "Français": ["ORL", "ASP", "Scanner TAP", "Uroscanner", "Échographie", "IRM", "CPRM", "CHC", "SA", "transsonore", "anéchogène", "Produit de contraste", "Voie Biliaire Principale", "Troncs Supra-Aortiques", "TDM", "TEP"],
    "English": ["ENT", "Plain abdominal X-ray", "CT CAP", "CT Urogram", "Ultrasound", "MRI", "MRCP", "HCC", "GA", "anechoic", "anechoic", "Contrast agent", "Common Bile Duct", "Supra-aortic trunks", "CT", "PET"],
    "Español": ["ORL", "Radiografía simple", "TC TAP", "Urotomografía", "Ecografía", "RM", "CPRM", "CHC", "Semanas", "anecoico", "anecoico", "Medio de contraste", "Vía Biliar Principal", "Troncos Supraaórticos", "TC", "PET"],
    "Deutsch": ["HNO", "Abdomenübersicht", "CT-TAB", "CT-Urographie", "Sonographie", "MRT", "MRCP", "HCC", "SSW", "anechoisch", "anechogen", "Kontrastmittel", "Hauptgallengang", "Supraaortale Äste", "CT", "PET"],
    "Italiano": ["ORL", "RX diretta addome", "TC TAP", "Uro-TC", "Ecografia", "RM", "CPRM", "HCC", "Settimane", "anecogeno", "anecogeno", "Mezzo di contrasto", "Via Biliare Principale", "Tronchi Sovraortici", "TC", "PET"],
    "Português": ["ORL", "Radiografia simples", "TC TAP", "Urotomografia", "Ecografia", "RM", "CPRM", "CHC", "IG", "anecoico", "anecoico", "Meio de contraste", "Via Biliar Principal", "Troncos Supra-aórticos", "TC", "PET"],
    "中文": ["耳鼻喉科", "腹部平片", "胸腹盆CT", "CT尿路造影", "超声", "MRI", "MRCP", "HCC", "孕周", "无回声", "无回声", "造影剂", "胆总管", "主动脉弓上干", "CT", "PET"],
    "日本語": ["耳鼻咽喉科", "腹部単純X線", "胸腹骨盤CT", "CTウログラフィー", "超音波", "MRI", "MRCP", "肝細胞癌", "妊娠週数", "無エコー", "無エコー", "造影剤", "総胆管", "大動脈弓上幹", "CT", "PET"],
    "한국어": ["이비인후과", "복부 단순 촬영", "흉복부 골반 CT", "CT 요로조영술", "초음파", "MRI", "MRCP", "간세포암", "임신 주수", "무에코", "무에코", "조영제", "총담관", "대동맥궁 상부 줄기", "CT", "PET"],
    "Русский": ["ЛОР", "Обзорная рентгенография", "КТ грудь/живот/таз", "КТ-урография", "УЗИ", "МРТ", "МРХПГ", "ГЦК", "Срок бер.", "анэхогенный", "анэхогенный", "Контрастное вещество", "ОЖП", "Супрааортальные стволы", "КТ", "ПЭТ"],
    "Türkçe": ["KBB", "Direkt batın grafisi", "Toraks-Batın-Pelvis BT", "BT Ürografi", "Ultrason", "MRG", "MRCP", "HCC", "Gebelik haftası", "anekoik", "anekoik", "Kontrast madde", "Ana Safra Kanalı", "Supra-aortik gövdeler", "BT", "PET"],
    "Nederlands": ["KNO", "Buikoverzichtsfoto", "CT Thorax-Abdomen", "CT-Urografie", "Echografie", "MRI", "MRCP", "HCC", "Zwangerschapsduur", "anachoïsch", "anechogeen", "Contrastmiddel", "Ductus Choledochus", "Supra-aortale vaten", "CT", "PET"],
    "Svenska": ["ÖNH", "Buköversikt", "DT Thorax-Buk", "DT Urografi", "Ultraljud", "MRT", "MRCP", "HCC", "Graviditetsvecka", "ekofri", "ekofri", "Kontrastmedel", "Gemensamma gallgången", "Supra-aortala kärl", "DT", "PET"],
    "Norsk": ["ØNH", "Oversikt abdomen", "CT Thorax-Abdomen", "CT Urografi", "Ultralyd", "MR", "MRCP", "HCC", "Svangerskapsuke", "ekkofritt", "ekkofritt", "Kontrastmiddel", "Hovedgallegang", "Supra-aortale kar", "CT", "PET"],
    "Dansk": ["ØNH", "Oversigt abdomen", "CT Thorax-Abdomen", "CT Urografi", "Ultralyd", "MR", "MRCP", "HCC", "Gestationsalder", "ekkofri", "ekkofri", "Kontraststof", "Den fælles galdegang", "Supra-aortale kar", "CT", "PET"],
    "Polski": ["Laryngologia", "RTG przeglądowe brzucha", "TK klatki/brzucha", "Urografia TK", "USG", "MRI", "MRCP", "HCC", "Tygodnie ciąży", "bezechowy", "bezechowy", "Środek kontrastowy", "Przewód żółciowy wspólny", "Pnie nadłukowe", "TK", "PET"],
    "Indonesien": ["THT", "Foto polos abdomen", "CT Thorax-Abdomen", "CT Urografi", "USG", "MRI", "MRCP", "HCC", "Usia kehamilan", "anekolik", "anekolik", "Media kontras", "Saluran Empedu Utama", "Trunkus Supra-Aorta", "CT", "PET"],
    "Thai": ["หู คอ จมูก", "เอกซเรย์ช่องท้อง", "CT ทรวงอก-ช่องท้อง", "CT ระบบทางเดินปัสสาวะ", "อัลตราซาวนด์", "MRI", "MRCP", "HCC", "อายุครรภ์", "ไม่สะท้อนเสียง", "ไม่สะท้อนเสียง", "สารทึบรังสี", "ท่อน้ำดีร่วม", "หลอดเลือดแดงใหญ่ส่วนบน", "CT", "PET"],
    "Malais": ["THT", "X-ray abdomen", "CT Thorax-Abdomen", "CT Urografi", "Ultrabunyi", "MRI", "MRCP", "HCC", "Minggu kehamilan", "anekoid", "anekoid", "Ejen kontras", "Saluran Hempedu Utama", "Trunkus Supra-Aorta", "CT", "PET"],
    "Grec": ["ΩΡΛ", "Ακτινογραφία κοιλίας", "Αξονική θώρακος-κοιλίας", "Αξονική ουρογραφία", "Υπερηχογράφημα", "Μαγνητική", "MRCP", "HCC", "Εβδομάδες κύησης", "ανηχοϊκή", "ανηχοϊκή", "Σκιαγραφικό", "Κοινός Χοληδόχος Πόρος", "Υπεραορτικά αγγεία", "Αξονική", "PET"],
    "Filipino": ["ENT", "Plain abdominal X-ray", "CT Chest-Abdomen", "CT Urogram", "Ultrasound", "MRI", "MRCP", "HCC", "Linggo ng pagbubuntis", "anechoic", "anechoic", "Contrast medium", "Common Bile Duct", "Supra-aortic trunks", "CT", "PET"],
    "Roumain": ["ORL", "Radiografie abdominală", "CT Torace-Abdomen", "Uro-CT", "Ecografie", "RMN", "Colangio-RMN", "HCC", "Săptămâni amenoree", "anecogen", "anecogen", "Substanță de contrast", "Calea Biliară Principală", "Trunchiuri Supra-Aortice", "CT", "PET"],
    "हिन्दी": ["ईएनटी", "पेट का एक्स-रे", "सीटी स्कैन छाती-पेट", "सीटी यूरोग्राम", "अल्ट्रासाउंड", "एमआरआई", "एमआरसीपी", "HCC", "गर्भावस्था की आयु", "एनेकोइक", "एनेकोइक", "कंट्रास्ट माध्यम", "सामान्य पित्त नली", "सुप्रा-एओर्टिक ट्रंक", "सीटी", "पीईटी"]
}

TARGET_LANGUAGES = {
    'Allemand': 'de', 'Anglais': 'en', 'Espagnol': 'es', 'Portugais': 'pt', 'Hindi': 'hi',
    'Suedois': 'sv', 'Turc': 'tr', 'Russe': 'ru', 'Neerlandais': 'nl', 'Norvegien': 'no',
    'Danois': 'da', 'Coreen': 'ko', 'Japonais': 'ja', 'Italien': 'it', 'Indonesien': 'id',
    'Polonais': 'pl', 'Thai': 'th', 'Malais': 'ms', 'Grec': 'el', 'Filipino': 'tl',
    'Chinois': 'zh-CN', 'Roumain': 'ro'
}

LANG_NAME_TO_DICT_KEY = {
    'Anglais': 'English', 'Chinois': '中文', 'Espagnol': 'Español', 'Allemand': 'Deutsch',
    'Italien': 'Italiano', 'Portugais': 'Português', 'Russe': 'Русский', 'Turc': 'Türkçe',
    'Suedois': 'Svenska', 'Norvegien': 'Norsk', 'Danois': 'Dansk', 'Neerlandais': 'Nederlands',
    'Japonais': '日本語', 'Coreen': '한국어', 'Indonesien': 'Indonesien', 'Polonais': 'Polski',
    'Thai': 'Thai', 'Malais': 'Malais', 'Grec': 'Grec', 'Filipino': 'Filipino',
    'Roumain': 'Roumain', 'Hindi': 'हिन्दी'
}


# SECTIONS CORRIGEES ET COMPLETES
RADIOLOGY_SECTIONS = {
    'Francais': { 
        'patterns': [
            r'(?i)(?:Indication(?:s)?|Renseignements\s+cliniques|Contexte)\s*[:\.]', 
            r'(?i)(?:Technique|Protocole|Modalité(?:s)?|Examen)\s*[:\.]', 
            r'(?i)(?:Résultats?|Interprétation|Revue\s+des\s+organes|Observations?)\s*[:\.]', 
            r'(?i)(?:Conclusions?|Synthèse|Impression)\s*[:\.]', 
            r'(?i)(?:Conduite\s+à\s+tenir|Recommandation(?:s)?|Propositions?)\s*[:\.]'
        ] 
    },
    'Allemand': { 
        'translations': ["Klinische Indikation:", "Bildgebungstechnik:", "Befunde:", "Beurteilung:", "Empfehlungen:"] 
    },
    'Anglais': { 
        'translations': ["Clinical indication:", "Imaging technique:", "Findings:", "Impression:", "Recommendations:"] 
    },
    'Espagnol': { 
        'translations': ["Indicación clínica:", "Técnica de imagen:", "Hallazgos:", "Impresión:", "Recomendaciones:"] 
    },
    'Italien': { 
        'translations': ["Indicazione clinica:", "Tecnica di imaging:", "Reperti:", "Conclusione:", "Raccomandazioni:"] 
    },
    'Portugais': { 
        'translations': ["Indicação clínica:", "Técnica de imagem:", "Achados:", "Impressão:", "Recomendações:"] 
    },
    'Chinois': { 
        'translations': ["临床指征：", "影像技术：", "影像所见：", "影像诊断：", "建议："] 
    },
    '日本語': { 
        'translations': ["臨床適応：", "検査手法：", "所見：", "診断：", "推奨事項："] 
    },
    'Russe': { 
        'translations': ["Клинические показания:", "Методика:", "Результаты:", "Заключение:", "Рекомендации:"] 
    },
    'Roumain': { 
        'translations': ["Indicație clinică:", "Tehnică imagistică:", "Rezultate:", "Concluzie:", "Recomandări:"] 
    },
    'Polski': { 
        'translations': ["Wskazania kliniczne:", "Technika badania:", "Wyniki:", "Wnioski:", "Zalecenia:"] 
    },
    'Turc': { 
        'translations': ["Klinik endikasyon:", "Görüntüleme tekniği:", "Bulgular:", "Sonuç:", "Öneriler:"] 
    },
    'Hindi': { 
        'translations': ["नैदानिक संकेत:", "इमेजिंग तकनीक:", "निष्कर्ष:", "प्रभाव:", "सिफारिशें:"] 
    },
    'Suedois': { 
        'translations': ["Klinisk indikation:", "Undersökningsmetodik:", "Fynd:", "Bedömning:", "Rekommendationer:"] 
    },
    'Neerlandais': { 
        'translations': ["Klinische indicatie:", "Onderzoekstechniek:", "Bevindingen:", "Conclusie:", "Aanbevelingen:"] 
    },
    'Norvegien': { 
        'translations': ["Klinisk indikasjon:", "Undersøkelsesteknikk:", "Funn:", "Konklusjon:", "Anbefalinger:"] 
    },
    'Danois': { 
        'translations': ["Klinisk indikation:", "Billedteknik:", "Fund:", "Konklusion:", "Anbefalinger:"] 
    },
    'Coreen': { 
        'translations': ["임상 적응증:", "검사 방법:", "영상 소견:", "결론:", "권고 사항:"] 
    },
    'Indonesien': { 
        'translations': ["Indikasi klinis:", "Teknik pencitraan:", "Temuan:", "Kesimpulan:", "Saran:"] 
    },
    'Thai': { 
        'translations': ["ข้อบ่งชี้ทางคลินิก:", "เทคนิคการตรวจ:", "ผลการตรวจ:", "สรุปผล:", "คำแนะนำ:"] 
    },
    'Malais': { 
        'translations': ["Indikasi klinikal:", "Teknik imej:", "Penemuan:", "Kesimpulan:", "Cadangan:"] 
    },
    'Grec': { 
        'translations': ["Κλινική ένδειξη:", "Τεχνική:", "Ευρήματα:", "Συμπέρασμα:", "Συστάσεις:"] 
    },
    'Filipino': { 
        'translations': ["Indikasyong klinikal:", "Teknik ng imaging:", "Resulta:", "Konklusyon:", "Rekomendasyon:"] 
    }
}


# ==========================================
# 2. LOGIQUE MÉTIER (WORKER)
# ==========================================

class TranslationWorker:
    def __init__(self, source_dir, output_dir, selected_langs, log_callback, progress_file_callback, progress_total_callback):
        self.source_dir = source_dir
        self.output_dir = output_dir
        self.selected_langs = selected_langs # Dict {name: code}
        self.log = log_callback
        self.update_file_progress = progress_file_callback
        self.update_total_progress = progress_total_callback
        self.stop_event = threading.Event()
        self.cache = self.load_cache()
        self.lock = threading.Lock()
        
        # OPIMISATION: Précompilation des expressions régulières
        self.protect_patterns = [
            (idx, re.compile(r'(?<=\d)' + re.escape(term) + r'\b|\b' + re.escape(term) + r'(?=\W|$)', re.IGNORECASE))
            for idx, term in enumerate(PROTECTED_TERMS)
        ]
        # SECURITE CLINIQUE: Sensible à la casse si le terme est un acronyme (tout en majuscule)
        self.gloss_patterns = []
        for idx, term in enumerate(KEYS_LIST):
            flags = 0 if term.isupper() else re.IGNORECASE
            self.gloss_patterns.append((idx, re.compile(r'\b' + re.escape(term) + r'\b', flags)))
            
        self.section_patterns = [
            (idx, re.compile(pat, re.IGNORECASE))
            for idx, pat in enumerate(RADIOLOGY_SECTIONS['Francais']['patterns'])
        ]

    # Patterns de détection/nettoyage de marqueurs corrompus (toutes versions, toutes langues)
    # [[ ZONE_0 ]], [[ ZONA_0 ]], [[ ゾーン_1 ]], [[ โซน_2 ]], [[ ЗОНА_3 ]], etc.
    _STALE_MARKER  = re.compile(r'\[\[\s*[^\]\[]{0,40}\d[^\]\[]{0,10}\]\]')
    _RESID_MARKER  = re.compile(r'⟦[^⟧]{0,20}⟧')   # ⟦s0⟧, ⟦p12⟧ … (V10 résiduel)

    @classmethod
    def _is_clean(cls, value):
        """Retourne True si la valeur ne contient aucun marqueur résiduel."""
        if not isinstance(value, str): return True
        return not (cls._STALE_MARKER.search(value) or cls._RESID_MARKER.search(value))

    @classmethod
    def _clean_cell(cls, text):
        """Supprime tous les marqueurs corrompus d'une cellule et nettoie les lignes vides résiduelles."""
        if not isinstance(text, str): return text
        text = cls._STALE_MARKER.sub('', text)
        text = cls._RESID_MARKER.sub('', text)
        # Supprimer les lignes devenues vides après suppression des marqueurs
        lines = [l for l in text.split('\n') if l.strip()]
        text = '\n'.join(lines)
        text = re.sub(r'[ \t]{2,}', ' ', text).strip()
        return text

    def load_cache(self):
        try:
            with open("pisum_cache.json", 'r', encoding='utf-8') as f:
                raw = json.load(f)
            # Purger les entrées corrompues (anciennes versions V9 ou V10 défectueux)
            clean = {k: v for k, v in raw.items() if self._is_clean(v)}
            purged = len(raw) - len(clean)
            if purged:
                print(f"[Cache] {purged} entrée(s) corrompue(s) purgée(s).")
            return clean
        except:
            return {}

    def save_cache(self):
        try:
            with open("pisum_cache.json", 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)
        except: pass

    def get_glossary_term(self, index, target_lang_name):
        dict_key = LANG_NAME_TO_DICT_KEY.get(target_lang_name, 'English')
        if dict_key in MEDICAL_GLOSSARY:
            terms = MEDICAL_GLOSSARY[dict_key]
            if index < len(terms): return terms[index]
        return MEDICAL_GLOSSARY['Français'][index]

    def google_translate_robust(self, text, target_lang):
        text_str = str(text).strip()
        if not text_str or len(text_str) < 2: return text
        
        # PAS de remplacement [[ -> (( : on garde les marqueurs intacts
        text_ready = text_str

        translator = GoogleTranslator(source='fr', target=target_lang)
        
        # Smart Chunking
        chunks = []
        limit = 4500
        temp_text = text_ready
        while len(temp_text) > limit:
            split_idx = max(temp_text.rfind('\n', 0, limit), temp_text.rfind('. ', 0, limit))
            if split_idx == -1: split_idx = limit
            chunks.append(temp_text[:split_idx+1])
            temp_text = temp_text[split_idx+1:]
        chunks.append(temp_text)

        translated_chunks = []
        for chunk in chunks:
            if self.stop_event.is_set(): return "STOPPED"
            
            # Si le chunk n'est que des symboles/chiffres, on le conserve tel quel
            if not any(c.isalpha() for c in chunk):
                translated_chunks.append(chunk)
                continue

            success = False
            for attempt in range(5): 
                try:
                    time.sleep(random.uniform(3.0, 6.0)) 
                    res = translator.translate(chunk)
                    if res:
                        translated_chunks.append(res)
                        success = True
                        break
                except Exception as e:
                    wait = (attempt + 1) * 3
                    self.log(f"⚠️ Erreur (Essai {attempt+1}/5) : {str(e)[:50]}... Pause {wait}s")
                    time.sleep(wait)
            
            if not success:
                self.log(f" ⚠️ Échec traduction bloc, conservation original.")
                translated_chunks.append(chunk)
        
        return "\n".join(translated_chunks)

    def translate_pipeline(self, text, target_lang_code, target_lang_name):
        if pd.isna(text) or str(text).strip() == "": return text
        text_str = str(text).strip()
        
        # 1. Filtre Rapide
        if not text_str: return text
        if len(re.findall(r'[a-zA-Z]', text_str)) < 2:
            return text
        
        # 2. Cache — vérifier que la valeur en cache est propre (pas de marqueur résiduel)
        cache_key = hashlib.sha256(f"{text_str}_{target_lang_code}".encode('utf-8')).hexdigest()
        if cache_key in self.cache:
            cached = self.cache[cache_key]
            if self._is_clean(cached):
                return cached
            else:
                # Valeur corrompue → supprimer et retraduire
                del self.cache[cache_key]

        # 3. Masking — Stratégie DEFINITIF avec marqueurs ⟦⟧ (U+27E6/U+27E7)
        # Ces caractères mathématiques sont TOTALEMENT opaques pour Google Translate :
        #   - Google ne les traduit jamais
        #   - Google ne les décompose pas
        #   - Les codes sont courts (s0..s4, p0..p99, g0..g13)
        # Restauration via restore_v2 qui tolère tout espacement interne résiduel.

        masked_text = text_str
        gloss_map   = {}   # '⟦g0⟧' -> glossary_index
        protect_map = {}   # '⟦p0⟧' -> original_term
        section_map = {}   # '⟦s0⟧' -> translated_section_label

        def _mk(prefix, idx):
            return f"⟦{prefix}{idx}⟧"

        def _restore(text, marker, replacement):
            """Restauration ultra-robuste : tolère espaces entre chaque caractère du marqueur."""
            inner = marker[1:-1]   # ex: 's0', 'p12', 'g3'
            spaced = r'\s*'.join(re.escape(c) for c in inner)
            pat = r'⟦\s*' + spaced + r'\s*⟧'
            # lambda évite que re.sub interprète les backslash dans replacement
            return re.sub(pat, lambda m: replacement, text, flags=re.IGNORECASE)

        # --- A. PROTECTION des termes médicaux invariants ---
        for idx, pat in self.protect_patterns:
            if pat.search(masked_text):
                marker = _mk('p', idx)
                protect_map[marker] = PROTECTED_TERMS[idx]
                masked_text = pat.sub(f" {marker} ", masked_text)

        # --- B. GLOSSAIRE médical ---
        for idx, pat in self.gloss_patterns:
            if pat.search(masked_text):
                marker = _mk('g', idx)
                gloss_map[marker] = idx
                masked_text = pat.sub(f" {marker} ", masked_text)

        # --- C. SECTIONS radiologiques ---
        french_patterns = self.section_patterns
        dict_key = LANG_NAME_TO_DICT_KEY.get(target_lang_name, 'English')
        if dict_key in RADIOLOGY_SECTIONS:
            target_sect_list = RADIOLOGY_SECTIONS[dict_key]['translations']
        elif target_lang_name in RADIOLOGY_SECTIONS:
            target_sect_list = RADIOLOGY_SECTIONS[target_lang_name]['translations']
        else:
            target_sect_list = RADIOLOGY_SECTIONS['Anglais']['translations']

        for idx, pat in french_patterns:
            matches = list(pat.finditer(masked_text))
            for match in reversed(matches):
                marker = _mk('s', idx)
                trans_sect = target_sect_list[idx] if idx < len(target_sect_list) else match.group(0)
                section_map[marker] = trans_sect
                masked_text = masked_text[:match.start()] + f" {marker} " + masked_text[match.end():]

        # 4. Traduction
        trans_text = self.google_translate_robust(masked_text, target_lang_code)
        if trans_text == "STOPPED": return text

        # 5. Restauration dans l'ordre : sections > glossaire > protected
        for marker, trans in section_map.items():
            trans_text = _restore(trans_text, marker, trans)

        for marker, idx in gloss_map.items():
            term_tr = self.get_glossary_term(idx, target_lang_name)
            trans_text = _restore(trans_text, marker, term_tr)

        for marker, val in protect_map.items():
            trans_text = _restore(trans_text, marker, val)

        # 6. Filet de sécurité : supprimer tout marqueur ⟦⟧ résiduel (V10)
        trans_text = re.sub(r'⟦[^⟧]{0,20}⟧', '', trans_text)

        # 7. Filet universel : supprimer tout marqueur [[ ... ]] résiduel (V9 et toutes langues)
        #    Couvre: [[ ZONE_0 ]], [[ ZONA_0 ]], [[ โซน_0 ]], [[ ゾーン_0 ]], etc.
        trans_text = re.sub(r'\[\[\s*[^\]]{0,30}\d\s*\]\]', '', trans_text)

        # 7.5 Nettoyage typographique clinique (Espaces autour de la ponctuation)
        trans_text = re.sub(r'\s+([.,;:!?)\]])', r'\1', trans_text)  # Supprime l'espace AVANT la ponctuation
        trans_text = re.sub(r'([(\[])\s+', r'\1', trans_text)        # Supprime l'espace APRÈS l'ouverture de parenthèse

        # 8. Nettoyage espaces multiples (sans toucher aux sauts de ligne)
        trans_text = re.sub(r'[ \t]{2,}', ' ', trans_text).strip()

        # Valider avant de mettre en cache
        if self._is_clean(trans_text):
            self.cache[cache_key] = trans_text
        return trans_text

    def format_excel(self, filepath):
        try:
            wb = load_workbook(filepath)
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=11)

            for sheet in wb.worksheets:
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                
                for cell in sheet[1]:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                
                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        cell.border = border

                for col in sheet.columns:
                    max_l = 0
                    col_l = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            max_l = max(max_l, max(len(l) for l in lines) if lines else 0)
                    sheet.column_dimensions[col_l].width = min(max_l + 2, 70)
            wb.save(filepath)
        except: pass


    def clean_sheet_name(self, name):
        if not name: return "Sheet1"
        clean = re.sub(r'[\\/*?:\[\]]', '_', str(name))
        return clean[:31]

    def process_language(self, l_name, l_code, data_dict, filename, current_op_ref, total_ops, stats):
        """Fonction qui gère UNE langue (exécutée en parallèle)"""
        target_folder = os.path.join(self.output_dir, l_name)
        os.makedirs(target_folder, exist_ok=True)
        target_file = os.path.join(target_folder, filename)

        if os.path.exists(target_file):
            self.log(f"  ⏭️ {l_name} : Déjà fait.")
            with self.lock:
                stats["skip"] += 1
                current_op_ref[0] += 1
                self.update_total_progress(current_op_ref[0], total_ops)
            return

        self.log(f"  🌍 Démarrage {l_name}...", same_line=False)
        
        try:
            processed_sheets = []
            local_data = {k: v.copy() for k, v in data_dict.items()}
            
            for sheet_name, df_target in local_data.items():
                raw_tr_sheet = self.google_translate_robust(sheet_name, l_code)
                tr_sheet = self.clean_sheet_name(raw_tr_sheet if raw_tr_sheet and raw_tr_sheet != "STOPPED" else sheet_name)
                
                df_target.columns = [self.translate_pipeline(c, l_code, l_name) for c in df_target.columns]
                
                for col in df_target.columns:
                    new_col = []
                    for val in df_target[col]:
                        if self.stop_event.is_set(): return
                        res = self.translate_pipeline(val, l_code, l_name)
                        new_col.append(res)
                    df_target[col] = new_col
                
                processed_sheets.append((tr_sheet, df_target))

            if processed_sheets and not self.stop_event.is_set():
                with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
                    for name, df in processed_sheets:
                        df.to_excel(writer, sheet_name=name, index=False)
                self.format_excel(target_file)
                self.log(f"  ✅ {l_name} Terminé !")
                with self.lock:
                    stats["ok"] += 1
        
        except Exception as e:
            self.log(f"  ❌ Erreur {l_name}: {e}")
            with self.lock:
                stats["error"] += 1
        finally:
            with self.lock:
                current_op_ref[0] += 1
                self.update_total_progress(current_op_ref[0], total_ops)

    def run(self):
        files = [f for f in os.listdir(self.source_dir) if f.endswith('.xlsx')]
        if not files:
            self.log("⚠️ Aucun fichier .xlsx trouvé dans le dossier source.")
            return

        # Créer automatiquement le dossier de sortie si absent
        os.makedirs(self.output_dir, exist_ok=True)

        total_ops = len(files) * len(self.selected_langs)
        current_op = [0]
        stats = {"ok": 0, "skip": 0, "error": 0}

        self.log(f"🚀 Démarrage OPTIMISÉ (Parallèle) — {len(files)} fichier(s) × {len(self.selected_langs)} langue(s)")

        for f_idx, filename in enumerate(files, 1):
            if self.stop_event.is_set(): break
            
            self.update_file_progress(f_idx, len(files))
            src_path = os.path.join(self.source_dir, filename)
            self.log(f"\n📄 Fichier : {filename} ({f_idx}/{len(files)})")
            
            try:
                xls = pd.ExcelFile(src_path)
                data_dict = {sheet: pd.read_excel(src_path, sheet_name=sheet) for sheet in xls.sheet_names}
            except Exception as e:
                self.log(f"❌ Erreur lecture: {e}")
                stats["error"] += 1
                continue

            with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
                futures = []
                for l_name, l_code in self.selected_langs.items():
                    if self.stop_event.is_set(): break
                    futures.append(
                        executor.submit(self.process_language, l_name, l_code, data_dict, filename, current_op, total_ops, stats)
                    )
                concurrent.futures.wait(futures)

            self.save_cache()

        self.log(f"\n{'='*50}")
        self.log(f"✨ TRAITEMENT TERMINÉ !")
        self.log(f"   ✅ Réussis  : {stats['ok']}")
        self.log(f"   ⏭️  Ignorés  : {stats['skip']} (déjà traduits)")
        self.log(f"   ❌ Erreurs  : {stats['error']}")
        self.log(f"{'='*50}")

# ==========================================
# 3. INTERFACE GRAPHIQUE (GUI)
# ==========================================

class PisumApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PISUM Translator Ultimate - Édition Radiologie")
        self.geometry("950x750")
        try: self.iconbitmap("pisum.ico") 
        except: pass
        
        self.worker = None
        self.source_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        self.create_widgets()
        
        self.source_path.set(os.getcwd())
        self.output_path.set(os.path.join(os.getcwd(), "TRADUCTIONS"))

    def create_widgets(self):
        # Frame Chemins
        frame_paths = ttk.LabelFrame(self, text="Configuration des Dossiers", padding=10)
        frame_paths.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_paths, text="Dossier Source (Excel Français) :").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_paths, textvariable=self.source_path, width=70).grid(row=0, column=1, padx=5)
        ttk.Button(frame_paths, text="Parcourir...", command=self.browse_source).grid(row=0, column=2)
        
        ttk.Label(frame_paths, text="Dossier de Sortie :").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame_paths, textvariable=self.output_path, width=70).grid(row=1, column=1, padx=5)
        ttk.Button(frame_paths, text="Parcourir...", command=self.browse_output).grid(row=1, column=2)

        # Frame Langues
        frame_langs = ttk.LabelFrame(self, text="Langues Cibles", padding=10)
        frame_langs.pack(fill="both", expand=True, padx=10, pady=5)
        
        canvas = tk.Canvas(frame_langs)
        scrollbar = ttk.Scrollbar(frame_langs, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.lang_vars = {}
        col = 0
        row = 0
        for name in sorted(TARGET_LANGUAGES.keys()):
            var = tk.BooleanVar(value=True)
            self.lang_vars[name] = var
            ttk.Checkbutton(self.scrollable_frame, text=name, variable=var).grid(row=row, column=col, sticky="w", padx=10, pady=2)
            col += 1
            if col > 3: 
                col = 0
                row += 1
        
        ttk.Button(frame_langs, text="Tout cocher", command=self.select_all).pack(side="left", padx=5, pady=5)
        ttk.Button(frame_langs, text="Tout décocher", command=self.deselect_all).pack(side="left", padx=5, pady=5)

        # Frame Progression
        frame_prog = ttk.LabelFrame(self, text="Progression", padding=10)
        frame_prog.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_prog, text="Fichier en cours :").grid(row=0, column=0, sticky="w")
        self.prog_file = ttk.Progressbar(frame_prog, orient="horizontal", length=600, mode="determinate")
        self.prog_file.grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(frame_prog, text="Total :").grid(row=1, column=0, sticky="w")
        self.prog_total = ttk.Progressbar(frame_prog, orient="horizontal", length=600, mode="determinate")
        self.prog_total.grid(row=1, column=1, padx=5, pady=2)

        # Frame Actions
        frame_actions = ttk.Frame(self, padding=10)
        frame_actions.pack(fill="both", expand=True, padx=10)
        
        self.btn_start = ttk.Button(frame_actions, text="▶ DÉMARRER LA TRADUCTION", command=self.start_process)
        self.btn_start.pack(side="top", fill="x", pady=5)
        
        self.btn_stop = ttk.Button(frame_actions, text="⏹ ARRÊTER", command=self.stop_process, state="disabled")
        self.btn_stop.pack(side="top", fill="x", pady=2)

        self.btn_purge = ttk.Button(frame_actions, text="🗑 PURGER LE CACHE (supprimer pisum_cache.json)", command=self.purge_cache)
        self.btn_purge.pack(side="top", fill="x", pady=2)

        self.btn_clean = ttk.Button(frame_actions, text="🧹 NETTOYER LES FICHIERS DÉJÀ TRADUITS (supprimer marqueurs [[...]])", command=self.clean_existing_files)
        self.btn_clean.pack(side="top", fill="x", pady=2)
        
        self.log_text = tk.Text(frame_actions, height=10, state="disabled", bg="#f0f0f0", font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, pady=5)

    def browse_source(self):
        d = filedialog.askdirectory()
        if d: self.source_path.set(d)

    def browse_output(self):
        d = filedialog.askdirectory()
        if d: self.output_path.set(d)

    def select_all(self):
        for v in self.lang_vars.values(): v.set(True)

    def deselect_all(self):
        for v in self.lang_vars.values(): v.set(False)

    def purge_cache(self):
        import os
        cache_file = "pisum_cache.json"
        if os.path.exists(cache_file):
            os.remove(cache_file)
            messagebox.showinfo("Cache purgé",
                "✅ pisum_cache.json supprimé.\n\n"
                "Tous les textes seront retraduits depuis zéro.\n"
                "Les marqueurs corrompus [[ ZONE_0 ]] disparaîtront définitivement.")
        else:
            messagebox.showinfo("Cache", "Aucun fichier cache trouvé (déjà propre).")

    def clean_existing_files(self):
        """Répare les fichiers xlsx déjà traduits en supprimant tous les marqueurs corrompus."""
        out = self.output_path.get()
        if not os.path.exists(out):
            messagebox.showerror("Erreur", "Dossier de sortie introuvable.")
            return
        if not messagebox.askyesno("Nettoyer les fichiers",
            f"Ceci va parcourir TOUS les fichiers .xlsx dans :\n{out}\n\n"
            "et supprimer les marqueurs corrompus [[...]] dans toutes les langues.\n\n"
            "Continuer ?"):
            return

        self.btn_start.config(state="disabled")
        self.btn_clean.config(state="disabled")
        self.log_text.config(bg="black", fg="white", state="normal")
        self.log_text.delete(1.0, "end")
        self.log_text.config(state="disabled")

        def _do_clean():
            total_files = 0
            total_cells = 0
            errors = 0
            for root, dirs, files in os.walk(out):
                for fname in files:
                    if not fname.endswith('.xlsx'): continue
                    fpath = os.path.join(root, fname)
                    try:
                        xls = pd.ExcelFile(fpath)
                        changed = False
                        sheets_data = {}
                        for sheet in xls.sheet_names:
                            df = pd.read_excel(fpath, sheet_name=sheet, dtype=str)
                            for col in df.columns:
                                for i, val in enumerate(df[col]):
                                    cleaned = TranslationWorker._clean_cell(val)
                                    if cleaned != val:
                                        df.at[i, col] = cleaned
                                        total_cells += 1
                                        changed = True
                            sheets_data[sheet] = df
                        if changed:
                            with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                                for sname, df in sheets_data.items():
                                    df.to_excel(writer, sheet_name=sname, index=False)
                            self.log(f"  ✅ {os.path.relpath(fpath, out)}")
                            total_files += 1
                    except Exception as e:
                        self.log(f"  ❌ {fname}: {e}")
                        errors += 1

            self.log(f"\n{'='*50}")
            self.log(f"🧹 NETTOYAGE TERMINÉ")
            self.log(f"   Fichiers modifiés : {total_files}")
            self.log(f"   Cellules nettoyées: {total_cells}")
            self.log(f"   Erreurs           : {errors}")
            self.log(f"{'='*50}")
            self.after(0, lambda: self.btn_start.config(state="normal"))
            self.after(0, lambda: self.btn_clean.config(state="normal"))
            self.after(0, lambda: self.log_text.config(bg="#f0f0f0", fg="black"))

        threading.Thread(target=_do_clean, daemon=True).start()

    def log(self, message, same_line=False):
        self.after(0, self._log_safe, message, same_line)

    def _log_safe(self, message, same_line):
        self.log_text.config(state="normal")
        if same_line: self.log_text.insert("end", message)
        else: self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def update_prog_file(self, current, total):
        self.after(0, self._update_prog_file_safe, current, total)

    def _update_prog_file_safe(self, current, total):
        self.prog_file["maximum"] = total
        self.prog_file["value"] = current

    def update_prog_total(self, current, total):
        self.after(0, self._update_prog_total_safe, current, total)

    def _update_prog_total_safe(self, current, total):
        self.prog_total["maximum"] = total
        self.prog_total["value"] = current

    def start_process(self):
        src = self.source_path.get()
        out = self.output_path.get()
        selected = {n: TARGET_LANGUAGES[n] for n, v in self.lang_vars.items() if v.get()}
        
        if not selected: return messagebox.showwarning("Attention", "Veuillez sélectionner au moins une langue.")
        if not os.path.exists(src): return messagebox.showerror("Erreur", "Dossier source introuvable.")

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.log_text.config(bg="black", fg="white") 
        self.log_text.delete(1.0, "end")
        
        self.worker = TranslationWorker(src, out, selected, self.log, self.update_prog_file, self.update_prog_total)
        thread = threading.Thread(target=self.run_worker_thread)
        thread.start()

    def run_worker_thread(self):
        self.worker.run()
        self.after(0, self.on_finished)

    def stop_process(self):
        if self.worker:
            self.worker.stop_event.set()
            self.log("\n⚠️ Arrêt en cours...")
            self.btn_stop.config(state="disabled")

    def on_finished(self):
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.log_text.config(bg="#f0f0f0", fg="black", state="normal")
        self.log_text.config(state="disabled")

if __name__ == "__main__":
    app = PisumApp()
    app.mainloop()